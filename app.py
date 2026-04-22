"""Data Center Water Risk Dashboard with two pages."""
import io

import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils import DEFAULT_DATA_PATH, inject_css, setup_sidebar

CITY_ORDER = ["Denver", "Miami", "Phoenix", "San Diego", "Seattle"]

CITY_PARAMS = {
    "Denver": {"electricity_rate": 0.133, "water_rate": 4.9, "range_c": 15, "coc": 4},
    "Miami": {"electricity_rate": 0.116, "water_rate": 5.5, "range_c": 20, "coc": 3},
    "Phoenix": {"electricity_rate": 0.131, "water_rate": 3.5, "range_c": 15, "coc": 3},
    "San Diego": {"electricity_rate": 0.295, "water_rate": 11.5, "range_c": 20, "coc": 3},
    "Seattle": {"electricity_rate": 0.119, "water_rate": 5.5, "range_c": 10, "coc": 4},
}

COOLING_DATA_POINTS = [
    "Electricity Rate ($/kWh)",
    "Total Electricity Cost ($/year)",
    "Load Factor (%)",
    "Range (°C)",
    "Tons (Cooling Load)",
    "Evap (gpm)",
    "Blowdown (gpm)",
    "Drift (gpm)",
    "Peak Makeup (gpm)",
    "Peak Makeup (MGD)",
    "Peak Annual Makeup (MGY)",
    "Actual Annual Makeup (MGY)",
    "Water Rate ($/1000 gal)",
    "Total Water Cost ($/year)",
]

DOLLAR_FIELDS = {
    "Electricity Rate ($/kWh)",
    "Total Electricity Cost ($/year)",
    "Water Rate ($/1000 gal)",
    "Total Water Cost ($/year)",
}

CITY_ACCENTS = {
    "Denver": "#FFFFFF",
    "Miami": "#CCCCCC",
    "Phoenix": "#AAAAAA",
    "San Diego": "#888888",
    "Seattle": "#555555",
}

st.set_page_config(
    page_title="Data Center Water Risk Dashboard",
    page_icon="💧",
    layout="wide",
    initial_sidebar_state="expanded",
)

inject_css()
setup_sidebar()

st.sidebar.markdown("<h1 style='color:#FFFFFF; font-family: Space Grotesk, sans-serif; font-weight:700; font-size:1.6rem; margin-bottom:0.25rem;'>Data Center Site Selection</h1>", unsafe_allow_html=True)
section = st.sidebar.radio(
    "Navigate",
    options=["Overview & User Guide", "Site Selection Analysis", "Environmental Impacts"],
)


def compute_city_metrics(city, dc_type, it_load):
    params = CITY_PARAMS[city]
    electricity_rate = params["electricity_rate"]
    water_rate = params["water_rate"]
    range_c = params["range_c"]
    coc = params["coc"]

    load_factor = 0.8 if dc_type == "AI/ML Cluster" else 0.6
    total_elec_cost = it_load * 1000 * 365 * electricity_rate
    tons = it_load * 1000 / 3.517
    evap = 0.001 * tons * range_c
    blowdown = evap / (coc - 1)
    drift = 0.01 * evap
    peak_makeup_gpm = evap + blowdown + drift
    peak_makeup_mgd = peak_makeup_gpm * 60 * 24 / 1_000_000
    peak_annual_mgy = peak_makeup_mgd * 365
    actual_annual_mgy = peak_annual_mgy * load_factor
    total_water_cost = (actual_annual_mgy * 1_000) * water_rate

    return {
        "Electricity Rate ($/kWh)": electricity_rate,
        "Total Electricity Cost ($/year)": total_elec_cost,
        "Load Factor (%)": load_factor * 100,
        "Range (°C)": range_c,
        "Tons (Cooling Load)": tons,
        "Evap (gpm)": evap,
        "Blowdown (gpm)": blowdown,
        "Drift (gpm)": drift,
        "Peak Makeup (gpm)": peak_makeup_gpm,
        "Peak Makeup (MGD)": peak_makeup_mgd,
        "Peak Annual Makeup (MGY)": peak_annual_mgy,
        "Actual Annual Makeup (MGY)": actual_annual_mgy,
        "Water Rate ($/1000 gal)": water_rate,
        "Total Water Cost ($/year)": total_water_cost,
    }


def fmt_value(field_name, value):
    if field_name in DOLLAR_FIELDS:
        return f"${value:,.2f}"
    return f"{value:,.2f}"


DOWNLOAD_BTN_CSS = """
<style>
div[data-testid="stDownloadButton"] button {
    background-color: #1A1A1A !important;
    border: 1px solid #2E2E2E !important;
    color: #FFFFFF !important;
}
</style>
"""


def _is_numeric_excel(val):
    if val is None or isinstance(val, bool):
        return False
    try:
        if pd.isna(val):
            return False
    except TypeError:
        pass
    try:
        float(val)
        return True
    except (TypeError, ValueError):
        return False


def _header_number_format(header):
    if "$" in header:
        return "$#,##0.00"
    if "(%)" in header:
        return "0.00%"
    return "#,##0.00"


def build_styled_xlsx_bytes(headers, rows, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    thin_white = Side(style="thin", color="FFFFFF")
    thin_gray = Side(style="thin", color="2E2E2E")
    border_header = Border(
        left=thin_white, right=thin_white, top=thin_white, bottom=thin_white
    )
    border_data = Border(
        left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray
    )

    header_fill = PatternFill(fill_type="solid", fgColor="FF0D0D0D")
    fill_odd = PatternFill(fill_type="solid", fgColor="FF1A1A1A")
    fill_even = PatternFill(fill_type="solid", fgColor="FF141414")

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11, color="FFF5F5F5")
    city_bold_font = Font(name="Calibri", size=11, bold=True, color="FFF5F5F5")

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_header
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 20

    for r, row_vals in enumerate(rows, start=2):
        is_odd = (r - 2) % 2 == 0
        row_fill = fill_odd if is_odd else fill_even
        ws.row_dimensions[r].height = 18
        for c, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c)
            header = headers[c - 1]
            cell.fill = row_fill
            cell.border = border_data
            cell.alignment = Alignment(vertical="center")
            if c == 1:
                cell.value = "" if val is None else str(val)
                cell.font = city_bold_font
                cell.number_format = "General"
            else:
                if _is_numeric_excel(val):
                    num = float(val)
                    if "(%)" in header:
                        cell.value = num / 100.0
                    else:
                        cell.value = num
                    cell.number_format = _header_number_format(header)
                    cell.font = data_font
                else:
                    cell.value = "" if pd.isna(val) else val
                    cell.number_format = "General"
                    cell.font = data_font

    for c in range(1, len(headers) + 1):
        col_letter = get_column_letter(c)
        max_len = len(str(headers[c - 1]))
        for r in range(2, 2 + len(rows)):
            v = ws.cell(row=r, column=c).value
            if v is not None and v != "":
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(40, max(15, max_len + 2))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


st.markdown("<h1 style='color:#FFFFFF; font-family: Space Grotesk, sans-serif; font-weight:700; font-size:2.5rem; margin-bottom: 0.25rem;'>Data Center Site Selection</h1>", unsafe_allow_html=True)
if section == "Overview & User Guide":
    st.title("Overview & User Guide")

    st.subheader("About")
    st.markdown(
        """
This dashboard is designed to support decision making in selecting suitable locations for data centers.
It allows the user to analyze key factors such as the cost of power and water, climate conditions, and
environmental impacts in a single, easy-to-use interface. Users can explore locations, adjust data center
configurations, and compare results to identify sites that best meet their needs. No technical background
is required — this tool is built to provide clear insights for confident decisions.
"""
    )

    st.subheader("Tips & Help")
    st.markdown("Use the table below as a quick reference guide for each part of the dashboard.")
    st.markdown(
        """
| App Text | Help Text |
| --- | --- |
| Select Cities | Select one or more cities from the dropdown list to compare. |
| Select Data Points to Compare | Use the dropdown list to select as many factors for comparison as required. See the definitions below for a description of each factor. |
| Data Center Type | Select the type of data center: AI/ML Cluster — Used for advanced computing (e.g., artificial intelligence models using GPUs) that requires more electricity for power and water for cooling. Hyperscale Cloud — Used for large-scale, general purpose cloud computing and storage (e.g., general web services using CPUs) that requires less electricity for power and water for cooling. This helps match locations to your needs. |
| IT Load | Enter the expected power demand for your data center up to 1,000 MW. Larger values mean more electricity is needed to power IT equipment and more water is needed for cooling. This input can affect which locations are suitable for the planned data center. |
"""
    )

    st.subheader("Site Selection Analysis — Data Point Definitions")
    st.markdown(
        """
- **Tons (Cooling Load):** This shows how much cooling the data center needs to remove heat from equipment. Higher values mean more cooling capacity is required.
- **Evap (gpm):** Water used during cooling as it evaporates to remove heat. This is the main source of water use in the system.
- **Blowdown (gpm):** Water that is discharged from the system to prevent buildup of minerals. This depends on water quality and system settings.
- **Drift (gpm):** A very small amount of water lost as fine mist during operation. This is typically minimal compared to other losses.
- **Peak Makeup (gpm):** The total water needed at peak operation, including all losses. This represents the maximum demand on the water supply at any moment.
- **Peak Makeup (MGD):** The highest total water required in a single day. This helps assess whether local water systems can meet demand.
- **Peak Annual Makeup (MGY):** The estimated maximum water use over a full year under peak conditions.
- **Actual Annual Makeup (MGY):** The expected total water use over a typical year, based on average operating conditions.
- **Water Rate ($/1000 gal):** The local cost of water. This directly affects operating expenses.
- **Total Water Cost ($/year):** The estimated yearly cost of water based on total usage and local rates.
"""
    )

    st.subheader("Environmental & Sustainability Metrics — Data Point Definitions")
    st.markdown(
        """
- **Carbon Price ($/metric ton):** The cost applied to each unit of carbon emissions. This value is used to estimate the environmental cost of energy use.
- **Annual Carbon Cost ($):** The estimated yearly cost of carbon emissions based on energy use and local carbon pricing.
- **Normalized Carbon Cost (Scale: 0–1):** A simplified score showing how carbon costs compare across locations. Lower values are better.
- **Climate Adjusted Water Use (gallons/year):** Estimated yearly water use adjusted for the local climate. Hotter or drier areas may require more water for cooling.
- **Raw Water Stress Index:** A measure of how limited or strained water resources are in this area. Higher values indicate greater pressure on water supply.
- **Calculated Water Stress Score (1=low, 5=high):** A simplified rating of water availability. Lower scores mean water is more readily available.
- **Sustainability Score:** An overall score combining environmental factors such as carbon impact and water use. Lower scores generally indicate a more sustainable location.
- **Monthly Precip (mm):** The average monthly rainfall in this location. Higher rainfall may help support water availability.
- **Total Impact (2=low, 11.5=high):** A combined score showing the overall environmental impact of the site. Lower values indicate more favorable conditions.
"""
    )
elif section == "Site Selection Analysis":
    st.title("Site Selection Analysis")

    selected_cities = st.multiselect(
        "Select Cities",
        options=CITY_ORDER,
        default=[],
    )

    selected_data_points = st.multiselect(
        "Select Data Points to Compare",
        options=COOLING_DATA_POINTS,
        default=["Total Electricity Cost ($/year)", "Total Water Cost ($/year)"],
    )

    city_inputs = {}
    if selected_cities:
        input_cols = st.columns(len(selected_cities))
        for idx, city in enumerate(selected_cities):
            with input_cols[idx]:
                st.markdown(f"**{city}**")
                dc_type = st.selectbox(
                    "Data Center Type",
                    options=["AI/ML Cluster", "Hyperscale Cloud"],
                    key=f"dc_type_{city}",
                )
                it_load = st.number_input(
                    "IT Load (MW)",
                    min_value=1,
                    max_value=1000,
                    value=100,
                    step=1,
                    key=f"it_load_{city}",
                )
                city_inputs[city] = {"dc_type": dc_type, "it_load": int(it_load)}

    city_results = {}
    for city in selected_cities:
        cfg = city_inputs.get(city, {"dc_type": "AI/ML Cluster", "it_load": 100})
        city_results[city] = compute_city_metrics(city, cfg["dc_type"], cfg["it_load"])

    st.markdown(
        """
        <style>
        .cooling-card {
            background-color: #141414;
            border: 1px solid #2E2E2E;
            border-radius: 12px;
            box-shadow: 0 2px 12px rgba(0, 0, 0, 0.4);
            padding: 24px;
            margin-bottom: 16px;
            overflow: hidden;
        }
        .cooling-card .accent-bar {
            height: 6px;
            border-radius: 10px;
            margin: -24px -24px 16px -24px;
        }
        .cooling-card .card-header {
            color: #F5F5F5;
            font-weight: 800;
            margin-bottom: 16px;
            padding-bottom: 10px;
            font-size: 1.5rem;
            border-bottom: 1px solid #FFFFFF;
        }
        .cooling-card .metric-row {
            display: flex;
            justify-content: space-between;
            align-items: baseline;
            gap: 12px;
            padding: 10px 0;
            border-bottom: 1px solid #2E2E2E;
        }
        .cooling-card .metric-row:last-child {
            border-bottom: none;
        }
        .cooling-card .metric-row.key-row {
            background: #1F1F1F;
            border-radius: 8px;
            padding: 12px 10px;
            margin: 4px 0;
        }
        .cooling-card .metric-label {
            color: #888888;
            font-size: 0.8rem;
            line-height: 1.2;
            max-width: 58%;
        }
        .cooling-card .metric-value {
            color: #FFFFFF;
            font-size: 1.4rem;
            font-weight: 800;
            text-align: right;
            line-height: 1.1;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    if selected_cities and selected_data_points:
        card_cols = st.columns(len(selected_cities))
        for idx, city in enumerate(selected_cities):
            accent = CITY_ACCENTS.get(city, "#CCCCCC")
            parts = [
                '<div class="cooling-card">',
                f'<div class="accent-bar" style="background-color: {accent};"></div>',
                f'<div class="card-header">{city}</div>',
            ]
            for metric in selected_data_points:
                val = city_results[city][metric]
                key_row_class = (
                    " key-row"
                    if metric in {"Total Electricity Cost ($/year)", "Total Water Cost ($/year)"}
                    else ""
                )
                parts.append(
                    (
                        f'<div class="metric-row{key_row_class}">'
                        f'<div class="metric-label">{metric}</div>'
                        f'<div class="metric-value">{fmt_value(metric, val)}</div>'
                        "</div>"
                    )
                )
            parts.append("</div>")
            with card_cols[idx]:
                st.markdown("".join(parts), unsafe_allow_html=True)
    elif selected_cities:
        st.info("Select at least one data point to display city cards.")

    if selected_cities:
        st.markdown(DOWNLOAD_BTN_CSS, unsafe_allow_html=True)
        headers_ss = ["City", "Data Center Type", "IT Load (MW)"] + list(
            selected_data_points
        )
        rows_ss = []
        for city in selected_cities:
            cfg = city_inputs.get(city, {"dc_type": "AI/ML Cluster", "it_load": 100})
            row = [city, cfg["dc_type"], cfg["it_load"]]
            for m in selected_data_points:
                row.append(city_results[city][m])
            rows_ss.append(row)
        xlsx_ss = build_styled_xlsx_bytes(
            headers_ss, rows_ss, "Site Selection Analysis"
        )
        st.download_button(
            label="⬇ Download to Excel",
            data=xlsx_ss,
            file_name="Site_Selection_Analysis.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_site_selection_xlsx",
        )

    chart_cities = selected_cities
    elec_vals = [city_results[city]["Total Electricity Cost ($/year)"] for city in chart_cities]
    water_vals = [city_results[city]["Total Water Cost ($/year)"] for city in chart_cities]

    fig_cc = go.Figure()
    fig_cc.add_trace(
        go.Bar(
            x=chart_cities,
            y=elec_vals,
            name="Total Electricity Cost ($/year)",
            marker_color="#FFFFFF",
            text=[f"${v:,.0f}" for v in elec_vals],
            textposition="outside",
            textfont=dict(color="#F5F5F5", size=11),
        )
    )
    fig_cc.add_trace(
        go.Bar(
            x=chart_cities,
            y=water_vals,
            name="Total Water Cost ($/year)",
            marker_color="#CCCCCC",
            text=[f"${v:,.0f}" for v in water_vals],
            textposition="outside",
            textfont=dict(color="#F5F5F5", size=11),
        )
    )
    fig_cc.update_layout(
        title=dict(text="Cooling Cost Comparison by City", font=dict(color="#F5F5F5", size=14)),
        barmode="group",
        template="plotly_white",
        paper_bgcolor="#141414",
        plot_bgcolor="#141414",
        font=dict(family="Space Grotesk, sans-serif", color="#F5F5F5", size=11),
        xaxis=dict(
            title="",
            tickangle=-30,
            showgrid=False,
            linecolor="#2E2E2E",
            tickfont=dict(color="#F5F5F5"),
        ),
        yaxis=dict(
            title="Cost ($/year)",
            showgrid=True,
            gridcolor="#2E2E2E",
            linecolor="#2E2E2E",
            tickfont=dict(color="#F5F5F5"),
        ),
        margin=dict(t=80, b=100),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1,
            font=dict(color="#F5F5F5"),
        ),
        height=500,
        hoverlabel=dict(
            bgcolor="#1A1A1A",
            font=dict(color="#F5F5F5", size=11),
            bordercolor="#2E2E2E",
        ),
    )
    fig_cc.update_yaxes(tickformat="$,.0f")
    st.plotly_chart(fig_cc, use_container_width=True)

else:
    st.title("Environmental Impacts")

    try:
        ENV_COLUMNS = [
            "City",
            "Avg Temp (°C)",
            "Avg Annual Precipitation (mm/year)",
            "Grid Carbon Intensity (kgCO2/kWh)",
            "Climate Adjusted Electricity (MWh/year)",
            "Annual CO2 (metric tons)",
            "Carbon Price ($/metric ton)",
            "Annual Carbon Cost ($)",
            "Normalized Carbon Cost (Scale: 0-1)",
            "Climate Adjusted Water Use (gallons/year)",
            "Raw Water Stress Index",
            "Calculated Water Stress Score (1=low, 5=high)",
            "Sustainability Score",
            "Monthly Precip (mm)",
            "Total Impact (2=low, 11.5=high)",
        ]

        df_env = pd.read_excel(
            DEFAULT_DATA_PATH,
            sheet_name="Cooling Cost Method",
            usecols="E:S",
            header=None,
            skiprows=17,
            nrows=5,
            engine="openpyxl",
        )
        df_env.columns = ENV_COLUMNS
        df_env = df_env.dropna(subset=["City"]).copy()
        df_env["City"] = df_env["City"].astype(str).str.strip()
        df_env = df_env[df_env["City"].isin(CITY_ORDER)].reset_index(drop=True)

        numeric_cols = [c for c in ENV_COLUMNS if c != "City"]
        for col in numeric_cols:
            df_env[col] = pd.to_numeric(df_env[col], errors="coerce")

        # Annual Carbon Cost
        df_env["Annual Carbon Cost ($)"] = (
            df_env["Annual CO2 (metric tons)"] * df_env["Carbon Price ($/metric ton)"]
        )

        # Normalized Carbon Cost (0-1 scale)
        L = df_env["Annual Carbon Cost ($)"]
        df_env["Normalized Carbon Cost (Scale: 0-1)"] = (
            (L - L.min()) / (L.max() - L.min())
        ).round(2)

        # Raw Water Stress Index
        df_env["Raw Water Stress Index"] = (
            df_env["Climate Adjusted Water Use (gallons/year)"]
            / df_env["Avg Annual Precipitation (mm/year)"]
        ).round(0)

        # Calculated Water Stress Score (1-5 scale)
        O = df_env["Raw Water Stress Index"]
        df_env["Calculated Water Stress Score (1=low, 5=high)"] = (
            1 + 4 * (O - O.min()) / (O.max() - O.min())
        ).round(0)

        # Total Impact Score
        df_env["Total Impact (2=low, 11.5=high)"] = (
            df_env["Calculated Water Stress Score (1=low, 5=high)"] * 2
            + df_env["Normalized Carbon Cost (Scale: 0-1)"] * 1.5
        )

        selected_cities_env = st.multiselect(
            "Select Cities",
            options=CITY_ORDER,
            default=[],
        )
        selected_data_points_env = st.multiselect(
            "Select Data Points to Compare",
            options=[
                "Avg Temp (°C)",
                "Avg Annual Precipitation (mm/year)",
                "Grid Carbon Intensity (kgCO2/kWh)",
                "Climate Adjusted Electricity (MWh/year)",
                "Annual CO2 (metric tons)",
                "Carbon Price ($/metric ton)",
                "Annual Carbon Cost ($)",
                "Normalized Carbon Cost (Scale: 0-1)",
                "Climate Adjusted Water Use (gallons/year)",
                "Raw Water Stress Index",
                "Calculated Water Stress Score (1=low, 5=high)",
                "Sustainability Score",
                "Monthly Precip (mm)",
                "Total Impact (2=low, 11.5=high)",
            ],
            default=[],
        )

        st.markdown(
            """
            <style>
            .env-card {
                background-color: #141414;
                border: 1px solid #2E2E2E;
                border-radius: 12px;
                box-shadow: 0 2px 12px rgba(0, 0, 0, 0.4);
                padding: 24px;
                margin-bottom: 16px;
            }
            .env-card .card-header {
                color: #F5F5F5;
                font-weight: 800;
                margin-bottom: 16px;
                padding-bottom: 10px;
                font-size: 1.5rem;
                border-bottom: 1px solid #FFFFFF;
            }
            .env-card .metric-row {
                display: flex;
                justify-content: space-between;
                align-items: baseline;
                gap: 12px;
                padding: 10px 0;
                border-bottom: 1px solid #2E2E2E;
            }
            .env-card .metric-row:last-child {
                border-bottom: none;
            }
            .env-card .metric-row.key-row {
                background: #1F1F1F;
                border-radius: 8px;
                padding: 12px 10px;
                margin: 4px 0;
            }
            .env-card .metric-label {
                color: #888888;
                font-size: 0.8rem;
                line-height: 1.2;
                max-width: 58%;
            }
            .env-card .metric-value {
                color: #FFFFFF;
                font-size: 1.4rem;
                font-weight: 800;
                text-align: right;
                line-height: 1.1;
            }
            </style>
            """,
            unsafe_allow_html=True,
        )

        if not selected_cities_env:
            st.info("Select one or more cities to view the Environmental Impacts dashboard.")
        else:
            env_df_selected = df_env[df_env["City"].isin(selected_cities_env)].copy()
            card_cols = st.columns(len(selected_cities_env))
            for idx, city in enumerate(selected_cities_env):
                city_row = env_df_selected[env_df_selected["City"] == city]
                if city_row.empty:
                    continue
                city_row = city_row.iloc[0]

                parts = [f'<div class="env-card"><div class="card-header">{city}</div>']
                for metric in selected_data_points_env:
                    value = city_row.get(metric)
                    value_fmt = f"{value:,.2f}" if pd.notna(value) else "N/A"
                    key_row_class = (
                        " key-row"
                        if metric
                        in {
                            "Total Impact (2=low, 11.5=high)",
                            "Calculated Water Stress Score (1=low, 5=high)",
                            "Sustainability Score",
                        }
                        else ""
                    )
                    parts.append(
                        (
                            f'<div class="metric-row{key_row_class}">'
                            f'<div class="metric-label">{metric}</div>'
                            f'<div class="metric-value">{value_fmt}</div>'
                            "</div>"
                        )
                    )
                parts.append("</div>")
                with card_cols[idx]:
                    st.markdown("".join(parts), unsafe_allow_html=True)

            st.markdown(DOWNLOAD_BTN_CSS, unsafe_allow_html=True)
            headers_env = ["City"] + list(selected_data_points_env)
            rows_env = []
            for city in selected_cities_env:
                city_row = env_df_selected[env_df_selected["City"] == city]
                if city_row.empty:
                    continue
                ser = city_row.iloc[0]
                row = [city]
                for m in selected_data_points_env:
                    row.append(ser.get(m))
                rows_env.append(row)
            xlsx_env = build_styled_xlsx_bytes(
                headers_env, rows_env, "Environmental Impacts"
            )
            st.download_button(
                label="⬇ Download to Excel",
                data=xlsx_env,
                file_name="Environmental_Impacts.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_environmental_impacts_xlsx",
            )
    except Exception as exc:
        st.error(f"Unable to load Environmental Impacts table: {exc}")
